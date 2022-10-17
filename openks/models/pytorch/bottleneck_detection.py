# Copyright (c) 2021 OpenKS Authors, Dlib Lab, Peking University. 
# All Rights Reserved.

import re
import os
import pandas as pd
from tqdm import tqdm

import pickle as pkl
from openpyxl import load_workbook
import numpy as np
from sentence_transformers import SentenceTransformer

from sknetwork.ranking import PageRank
import sknetwork

from ..model import BottleneckModel
from .relavance_modules import RelavanceScore

#%%

class DataSet(object):

    # load data from a .xlsx file;
    # invoke DataSet.load_data() to get : 
    # China_id,
    # concept_list, 
    # kg_paper_concept,
    # kg_paper_country, 
    # kg_paper_paper,
    # paper_name_embeddings,
    # concept_name_embeddings;
    # or you can load cached data directly, it's much faster

    def __init__(self, path = 'relavance_modules/论文引用网络数据集.xlsx'):
        self.path = path
        tmp = path.split('/')
        self.prefix = '/'.join(tmp[:-1]) + '/'

    def load_data(self):

        print('loading xlsx file...')
        wb = load_workbook(self.path)
        sheets = wb.get_sheet_names()
        sh = wb.get_sheet_by_name(sheets[0])
        rows = sh.rows
        columns = sh.columns

        concept = wb.get_sheet_by_name(sheets[0]) #391 id name name_en desc level keyword_zh keyword_en 1,2:name
        concept_hierarchy = wb.get_sheet_by_name(sheets[1]) #391 id name parent_id parent_name 1,2:name
        paper = wb.get_sheet_by_name(sheets[2]) #190160 id doi title year 1,2:doi
        citation = wb.get_sheet_by_name(sheets[3]) #770273 article_id title year refer_id refer_title refer_year 1,2:title
        paper_author = wb.get_sheet_by_name(sheets[4]) #496481 id title expert_id expert_name 1,2:title
        paper_concept = wb.get_sheet_by_name(sheets[5]) #208119 id title node_id node_name 1,2:title
        author_affiliation = wb.get_sheet_by_name(sheets[6]) #233529 expert_id expert_name org resume 1,2:expert_name

        ####################################################################
        # map affiliation to integer
        affiliation_set = set()
        affiliation_list = []
        if os.path.exists(self.prefix + 'kg_affiliation_list'):
            with open(self.prefix + 'kg_affiliation_list', 'rb') as fl:
                affiliation_list = pkl.load(fl)
        else:
            print('mapping affiliation to integer ... ')
            L = len(list(author_affiliation))
            for i in tqdm(range(2,L,1)):
                expert_id = author_affiliation.cell(row=i,column=1).value
                org = author_affiliation.cell(row=i,column=3).value
                if org is None:
                    continue
                affiliation_set.add(org)
            affiliation_list = list(affiliation_set)
            with open(self.prefix + 'kg_affiliation_list', 'wb') as fl:
                pkl.dump(affiliation_list, fl)

        ####################################################################
        # map author to integer
        author_id = {}
        if os.path.exists(self.prefix + 'kg_author_id'):
            with open(self.prefix + 'kg_author_id', 'rb') as fl:
                author_id = pkl.load(fl)
        else:
            print('mapping author to integer')
            L = len(list(author_affiliation))
            author_set = set()
            for i in tqdm(range(2,L,1)):
                expert_id = author_affiliation.cell(row=i,column=1).value
                author_set.add(expert_id)
            author_list = list(author_set)
            author_id  = dict(list(zip(author_list, list(range(len(author_list))))))
            with open(self.prefix + 'kg_author_id', 'wb') as fl:
                pkl.dump(author_id, fl)
        # print(list(author_id.keys())[:10])

        ####################################################################
        # map paper to integer
        paper_id = {}
        if os.path.exists(self.prefix + 'kg_paper_id'):
            with open(self.prefix + 'kg_paper_id', 'rb') as fl:
                paper_id = pkl.load(fl)
        else:
            print('mapping paper to integer...')
            L = len(list(paper))
            for i in tqdm(range(2, L, 1)):
                p_id = paper.cell(row=i,column=1).value
                if p_id in paper_id.keys():
                    print(p_id)
                    raise Exception
                paper_id[p_id] = i-2
            with open(self.prefix + 'kg_paper_id', 'wb') as fl:
                pkl.dump(paper_id, fl)
        # print(min(list(paper_id.values())))
        L = len(list(paper))
        paper_name_list = []
        for i in range(2, L, 1):
            p_id = paper.cell(row=i,column=1).value
            nm = paper.cell(row=i,column=3).value
            paper_name_list.append(nm)
            if paper_id[p_id] != i-2:
                raise Exception

        ####################################################################
        # map concept to integer, but ingnore concepts those are not leaf nodes
        L = len(list(concept))
        # print(L)
        # print(len(list(concept_hierarchy)))
        p_set = set()
        for i in range(2, L, 1):
            parent_id = concept_hierarchy.cell(row=i, column=3).value
            p_set.add(parent_id)
        A = 0
        Gid = 0
        concept_id = {}
        true_concept = []
        for i in range(2,L,1):
            c_id = concept.cell(row=i, column=1).value
            if c_id not in p_set:
                concept_id.update({c_id:A})
                if concept.cell(row=i, column=3).value == 'Model reasoning':
                    Gid = A
                    # print('*', A)
                true_concept.append(concept.cell(row=i, column=3).value)
                A += 1        
        # print(A, Gid)
        with open(self.prefix+'concept.txt', 'w') as fl:
            for i,c in enumerate(true_concept):
                fl.write(str(i)+','+str(c)+'\n')

        ####################################################################
        # get country name
        add = 0
        country = []
        with open(self.prefix +'country.txt', 'r') as fl:
            for txt in fl.readlines():
                if('-' in txt):
                    txt_list=txt.split('-')
                    short_name = txt_list[0]
                    full_name = re.findall(r'\((.*)\)', txt_list[1])[0].lower()
                    tmp = re.findall(r'(.*) *\(.*\) *', full_name)
                    if (len(tmp)>0):
                        full_name = tmp[0]
                    country.append((short_name, full_name))
        China_id = 0
        for i,(n1,n2) in enumerate(country):
            if 'china' == n2:
                China_id = i
                break
        print('China_id:', China_id)

        ####################################################################
        print('aligning affiliation and country...')
        affiliation_country = {}
        for nm in tqdm(affiliation_list):
            if isinstance(nm,str):
                bo = False
                for i,(n1,n2) in enumerate(country) :
                    if(n2 in nm.lower()):
                        affiliation_country.update({nm:i})
                        bo = True
                        break
                    else:
                        Cname = nm.split(',')[-1]
                        Cname = re.findall(r' *(.*)', Cname)[0]
                        L = len(n1)
                        if(n1 == Cname[:L]):
                            affiliation_country.update({nm:i})
                            bo = True
                            break
                if bo is not True:
                    if 'national' in nm.lower():
                        affiliation_country.update({nm:China_id})
            else:
                pass

        ####################################################################
        author_country = {}
        if os.path.exists(self.prefix + 'kg_author_country'):
            with open(self.prefix + 'kg_author_country', 'rb') as fl:
                author_country = pkl.load(fl)
        else:
            print('aligning author and country...')

            L = len(list(author_affiliation))
            for i in tqdm(range(2,L,1)):
                expert_id = author_affiliation.cell(row=i,column=1).value
                org = author_affiliation.cell(row=i,column=3).value
                if org is None:
                    continue
                if org in affiliation_country.keys():
                    a_id = author_id[expert_id]
                    if a_id not in author_country.keys():
                        author_country[a_id] = []
                    author_country[a_id].append(affiliation_country[org])
            for k,p in author_country.items():
                author_country[k] = list(set(p))
            with open(self.prefix + 'kg_author_country', 'wb') as fl:
                pkl.dump(author_country, fl)

        ####################################################################
        paper_country = {}
        if os.path.exists(self.prefix +'kg_paper_country'):
            with open(self.prefix +'kg_paper_country', 'rb') as fl:
                paper_country = pkl.load(fl)
        else:
            print('aligning paper and country...')
            L = len(list(paper_author))
            for i in tqdm(range(2, L, 1)):
                p_id = paper_author.cell(row=i,column=1).value
                e_id = paper_author.cell(row=i, column=3).value
                if e_id not in author_id.keys():
                    continue
                p_id = paper_id[p_id]
                e_id = author_id[e_id]
                if p_id in paper_country.keys():
                    continue
                if e_id in author_country.keys():
                    paper_country[p_id] = author_country[e_id]
            with open(self.prefix +'kg_paper_country', 'wb') as fl:
                pkl.dump(paper_country, fl)

        ####################################################################
        paper_concept_d = {}
        # weird_L = set()
        if os.path.exists(self.prefix +'kg_paper_concept'):
            with open(self.prefix +'kg_paper_concept', 'rb') as fl:
                paper_concept_d = pkl.load(fl)
        else:
            print('aligning paper and concept...')
            L = len(list(paper_concept))
            for i in tqdm(range(2,L,1)):
                p_id = paper_concept.cell(row=i,column=1).value
                c_id = paper_concept.cell(row=i,column=3).value
                if c_id not in concept_id.keys():
                    continue
                c_id = concept_id[c_id]
                # if(c_id == Gid):
                #     weird_L.add(str(p_id))
                p_id = paper_id[p_id]
                if p_id not in paper_concept_d.keys():
                    paper_concept_d[p_id] = set()
                paper_concept_d[p_id].add(c_id)
            with open(self.prefix +'kg_paper_concept', 'wb') as fl:
                pkl.dump(paper_concept_d, fl)

            
        ####################################################################
        paper_paper = {}
        if os.path.exists(self.prefix + 'kg_paper_paper'):
            with open(self.prefix + 'kg_paper_paper', 'rb') as fl:
                paper_paper = pkl.load(fl)
        else:
            print('aligning paper and paper...')
            L = len(list(citation))
            for i in tqdm(range(2,L,1)):
                f_id = citation.cell(row=i,column=1).value
                t_id = citation.cell(row=i,column=4).value
                f_id = paper_id[f_id]
                t_id = paper_id[t_id]
                if f_id not in paper_paper.keys():
                    paper_paper[f_id] = []
                paper_paper[f_id].append(t_id)
            with open(self.prefix + 'kg_paper_paper', 'wb') as fl:
                pkl.dump(paper_paper, fl)

        def load_embeddings(file_name, text):

            if os.path.exists(file_name):
                with open(file_name, 'rb') as fl:
                    ret = pkl.load(fl)
            else:
                print('getting embeddings for '+file_name)
                model = SentenceTransformer('all-mpnet-base-v2')
                ret = model.encode(text)
                with open(file_name, 'wb') as fl:
                    pkl.dump(ret, fl)
            return ret

        paper_name_embeddings = load_embeddings(self.prefix + 'paper_name_embeddings', paper_name_list)
        concept_name_embeddings = load_embeddings(self.prefix + 'concept_name_embeddings', true_concept)
        
        return China_id, \
            true_concept, \
            paper_concept_d, \
            paper_country, \
            paper_paper, \
            paper_name_embeddings, \
            concept_name_embeddings


@BottleneckModel.register("Bottleneck", "PyTorch")
class BottleneckDetection(BottleneckModel):
    
    def __init__(self, name='pytorch-default', path= 'relavance_modules/论文引用网络数据集.xlsx'):
        self.name = name
        self.path = path
        
    def get_data(self):
        A = DataSet(self.path)
        return A.load_data()
    
    def run(self):
        China_id, true_concept, paper_concept, paper_country, paper_paper, \
        paper_name_embeddings, concept_name_embeddings = self.get_data()
        paper_id, country_id, concept_id = {}
        for p in paper_concept:
            if p not in paper_id:
                paper_id[p] = len(paper_id)
            for c in paper_concept[p]:
                if c not in concept_id:
                    concept_id[c] = len(concept_id)
        for p in paper_paper:
            if p not in paper_id:
                paper_id[p] = len(paper_id)
            for c in paper_paper[p]:
                if c not in paper_id:
                    paper_id[c] = len(paper_id)
        for p in paper_country:
            if p not in paper_id:
                paper_id[p] = len(paper_id)
            for c in paper_country[p]:
                if c not in country_id:
                    country_id[c] = len(country_id)
        for c in concept_id:
            concept_id[c] += len(paper_id)
        for c in country_id:
            country_id[c] += len(paper_id) + len(concept_id)
        edge_list = []
        for p in paper_concept:
            for c in paper_concept[p]:
                edge_list.append((paper_id[p], concept_id[c], 1)) 
        for p in paper_paper:
            for c in paper_paper[p]:
                edge_list.append((paper_id[p], paper_id[c], 1)) 
        for p in paper_country:
            for c in paper_country[p]:
                edge_list.append((paper_id[p], country_id[c], 1))
        adj = sknetwork.utils.edgelist2adjacency(edge_list, undirected=True)
        seeds = {0: 1}
        relavance_score = RelavanceScore()
        total = relavance_score.run(adj, seeds, paper_id, country_id, concept_id, paper_country, paper_concept)
        country_sum = total.toarray().sum(0)
        china_id_new = country_id[China_id] -(len(paper_id) + len(concept_id))
        china = total.toarray()[china_id_new]
        concept_china = {}
        for i in range(len(concept_id)):
            if country_sum[i] == 0:
                concept_china[i] = 0
                continue
            concept_china[i] = china[i] / country_sum[i]
        id_concept = {}
        for c in concept_id:
            id_concept[concept_id[c] - (len(paper_id))] = c
        bottlenecks = []
        concept_score = {}
        for c in concept_china:
            if concept_china[c] == 0:
                bottlenecks.append(true_concepts[id_concept[c]])
        for c in concept_china:
            concept_score[true_concepts[id_concept[c]]] = concept_china[c]
        return bottlenecks, concept_score
